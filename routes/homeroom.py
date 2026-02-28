from flask import Blueprint, request, jsonify, send_file, session
from utils.db import get_db_connection, sanitize_input, sanitize_html
import io

homeroom_bp = Blueprint('homeroom', __name__)

# ============================================
# 담임 자격 확인 API
# ============================================
@homeroom_bp.route('/api/homeroom/check', methods=['GET'])
def check_homeroom_teacher():
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        member_id = sanitize_input(request.args.get('member_id'), 50)
        
        if not member_id:
            return jsonify({'success': False, 'message': '회원 정보가 필요합니다.'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        
        cursor = conn.cursor()
        
        if school_id:
            cursor.execute("""
                SELECT class_grade, class_no, member_name FROM tea_all 
                WHERE school_id = %s AND member_id = %s AND class_grade IS NOT NULL AND class_no IS NOT NULL
            """, (school_id, member_id))
        else:
            cursor.execute("""
                SELECT class_grade, class_no, member_name FROM tea_all 
                WHERE member_school = %s AND member_id = %s AND class_grade IS NOT NULL AND class_no IS NOT NULL
            """, (member_school, member_id))
        
        result = cursor.fetchone()
        
        if result and result['class_grade'] and result['class_no']:
            return jsonify({
                'success': True, 
                'is_homeroom': True,
                'class_grade': result['class_grade'],
                'class_no': result['class_no'],
                'teacher_name': result['member_name']
            })
        else:
            return jsonify({'success': True, 'is_homeroom': False, 'message': '담임 학급이 배정되지 않았습니다.'})
        
    except Exception as e:
        print(f"담임 확인 오류: {e}")
        return jsonify({'success': False, 'message': '담임 확인 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 담임 교사 정보 조회 API
# ============================================
@homeroom_bp.route('/api/homeroom/teacher', methods=['GET'])
def get_homeroom_teacher():
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        class_grade = sanitize_input(request.args.get('class_grade'), 10)
        class_no = sanitize_input(request.args.get('class_no'), 10)
        
        if not class_grade or not class_no:
            return jsonify({'success': False, 'message': '학급 정보가 필요합니다.'})
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        
        cursor = conn.cursor()
        
        if school_id:
            cursor.execute("""
                SELECT member_name, member_id FROM tea_all 
                WHERE school_id = %s AND class_grade = %s AND class_no = %s LIMIT 1
            """, (school_id, class_grade, class_no))
        else:
            cursor.execute("""
                SELECT member_name, member_id FROM tea_all 
                WHERE member_school = %s AND class_grade = %s AND class_no = %s LIMIT 1
            """, (member_school, class_grade, class_no))
        
        result = cursor.fetchone()
        
        if result:
            return jsonify({'success': True, 'teacher_name': result['member_name'], 'teacher_id': result['member_id']})
        else:
            return jsonify({'success': True, 'teacher_name': None, 'message': '담임 선생님 정보가 없습니다.'})
        
    except Exception as e:
        print(f"담임 정보 조회 오류: {e}")
        return jsonify({'success': False, 'message': '담임 정보 조회 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학급 학생 목록 조회 API
# ============================================
@homeroom_bp.route('/api/homeroom/students', methods=['GET'])
def get_homeroom_students():
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        class_grade = sanitize_input(request.args.get('class_grade'), 10)
        class_no = sanitize_input(request.args.get('class_no'), 10)

        if not class_grade or not class_no:
            return jsonify({'success': False, 'message': '학급 정보가 필요합니다.'})

        # [보안] 학생/학부모: 본인 반만 조회 가능
        user_role = session.get('user_role')
        if user_role in ('student', 'parent'):
            if class_grade != session.get('class_grade') or class_no != session.get('class_no'):
                return jsonify({'success': False, 'message': '본인 학급만 조회할 수 있습니다.'}), 403
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        
        cursor = conn.cursor()
        
        if school_id:
            cursor.execute("""
                SELECT id, member_id, member_name, member_birth, member_tel, class_num, class_role, point
                FROM stu_all WHERE school_id = %s AND class_grade = %s AND class_no = %s ORDER BY CAST(class_num AS UNSIGNED) ASC, class_num ASC
            """, (school_id, class_grade, class_no))
        else:
            cursor.execute("""
                SELECT id, member_id, member_name, member_birth, member_tel, class_num, class_role, point
                FROM stu_all WHERE member_school = %s AND class_grade = %s AND class_no = %s ORDER BY CAST(class_num AS UNSIGNED) ASC, class_num ASC
            """, (member_school, class_grade, class_no))

        students = cursor.fetchall()
        student_list = [{'id': s['id'], 'member_id': s['member_id'], 'member_name': s['member_name'],
                         'member_birth': str(s['member_birth']) if s['member_birth'] else '',
                         'member_tel': s['member_tel'],
                         'class_num': s['class_num'], 'class_role': s['class_role'] or '', 'point': s['point']} for s in students]
        
        return jsonify({'success': True, 'students': student_list, 'count': len(student_list)})
        
    except Exception as e:
        print(f"학급 학생 조회 오류: {e}")
        return jsonify({'success': False, 'message': '학생 조회 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학생 검색 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/search', methods=['GET'])
def search_school_students():
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        class_grade = sanitize_input(request.args.get('class_grade'), 10)
        class_no = sanitize_input(request.args.get('class_no'), 10)
        keyword = sanitize_input(request.args.get('keyword'), 50)

        if not keyword or len(keyword.strip()) < 1:
            return jsonify({'success': False, 'message': '검색어를 입력해주세요.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()

        if school_id:
            cursor.execute("""
                SELECT id, member_id, member_name, member_birth, member_tel, class_grade, class_no, class_num
                FROM stu_all WHERE school_id = %s AND member_name LIKE %s
                AND NOT (class_grade <=> %s AND class_no <=> %s) ORDER BY member_name ASC LIMIT 20
            """, (school_id, f'%{keyword}%', class_grade, class_no))
        else:
            cursor.execute("""
                SELECT id, member_id, member_name, member_birth, member_tel, class_grade, class_no, class_num
                FROM stu_all WHERE member_school = %s AND member_name LIKE %s
                AND NOT (class_grade <=> %s AND class_no <=> %s) ORDER BY member_name ASC LIMIT 20
            """, (member_school, f'%{keyword}%', class_grade, class_no))

        students = cursor.fetchall()
        result = [{'id': s['id'], 'member_id': s['member_id'], 'member_name': s['member_name'],
                   'member_birth': str(s['member_birth']) if s['member_birth'] else '',
                   'member_tel': s['member_tel'],
                   'class_grade': s['class_grade'], 'class_no': s['class_no'], 'class_num': s['class_num']} for s in students]

        return jsonify({'success': True, 'students': result})

    except Exception as e:
        print(f"학생 검색 오류: {e}")
        return jsonify({'success': False, 'message': '검색 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학생 배정 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/assign', methods=['POST'])
def assign_student_to_class():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        student_db_id = data.get('id')
        class_grade = sanitize_input(data.get('class_grade'), 10)
        class_no = sanitize_input(data.get('class_no'), 10)
        class_num = sanitize_input(data.get('class_num'), 10)

        if not student_db_id or not class_grade or not class_no:
            return jsonify({'success': False, 'message': '필수 정보가 부족합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()
        cursor.execute("UPDATE stu_all SET class_grade = %s, class_no = %s, class_num = %s, updated_at = NOW() WHERE id = %s",
                        (class_grade, class_no, class_num, student_db_id))
        conn.commit()
        return jsonify({'success': True, 'message': '학생이 반에 배정되었습니다.'})

    except Exception as e:
        print(f"학생 배정 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '배정 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학생 추가 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/add', methods=['POST'])
def add_student():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        school_id = sanitize_input(data.get('school_id'), 50)
        member_school = sanitize_input(data.get('member_school'), 100)
        member_name = sanitize_input(data.get('member_name'), 100)
        member_birth = sanitize_input(data.get('member_birth'), 20)
        member_tel = sanitize_input(data.get('member_tel'), 20)
        class_grade = sanitize_input(data.get('class_grade'), 10)
        class_no = sanitize_input(data.get('class_no'), 10)
        class_num = sanitize_input(data.get('class_num'), 10)

        if not member_name:
            return jsonify({'success': False, 'message': '학생 이름은 필수입니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()

        if school_id:
            cursor.execute("SELECT id, member_name, member_birth, member_tel FROM stu_all WHERE school_id = %s AND member_name = %s", (school_id, member_name))
        else:
            cursor.execute("SELECT id, member_name, member_birth, member_tel FROM stu_all WHERE member_school = %s AND member_name = %s", (member_school, member_name))

        candidates = cursor.fetchall()
        existing = None
        for c in candidates:
            tel_match = (member_tel and c['member_tel'] and member_tel == c['member_tel'])
            birth_match = (member_birth and c['member_birth'] and member_birth == c['member_birth'])
            if tel_match or birth_match:
                existing = c
                break

        if existing:
            update_parts = ["class_grade = %s", "class_no = %s", "updated_at = NOW()"]
            update_params = [class_grade, class_no]
            if class_num:
                update_parts.append("class_num = %s"); update_params.append(class_num)
            if member_tel and not existing['member_tel']:
                update_parts.append("member_tel = %s"); update_params.append(member_tel)
            if member_birth and not existing['member_birth']:
                update_parts.append("member_birth = %s"); update_params.append(member_birth)
            update_params.append(existing['id'])
            cursor.execute(f"UPDATE stu_all SET {', '.join(update_parts)} WHERE id = %s", update_params)
            conn.commit()
            return jsonify({'success': True, 'message': f'{member_name} 학생이 반에 배정되었습니다. (기존 학생)', 'is_new': False})
        else:
            cursor.execute("""
                INSERT INTO stu_all (member_id, school_id, member_school, member_name, member_birth, member_tel, class_grade, class_no, class_num, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            """, ('', school_id, member_school, member_name, member_birth or None, member_tel or None, class_grade, class_no, class_num or None))
            conn.commit()
            return jsonify({'success': True, 'message': f'{member_name} 학생이 새로 등록되었습니다.', 'is_new': True})

    except Exception as e:
        print(f"학생 추가 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '학생 추가 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학생 제거 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/remove', methods=['POST'])
def remove_student_from_class():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        student_db_id = data.get('id')
        if not student_db_id:
            return jsonify({'success': False, 'message': 'ID가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()
        cursor.execute("UPDATE stu_all SET class_grade = NULL, class_no = NULL, class_num = NULL, updated_at = NOW() WHERE id = %s", (student_db_id,))
        conn.commit()
        return jsonify({'success': True, 'message': '학생이 반에서 제거되었습니다.'})

    except Exception as e:
        print(f"학생 제거 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '제거 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학생 엑셀 템플릿 다운로드 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/template', methods=['GET'])
def download_student_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '학생명단'

        header_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='2563EB')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:F1')
        ws['A1'] = '※ 이름은 필수입니다. 나머지는 있는 항목만 입력하세요. 번호는 숫자만 입력하세요.'
        ws['A1'].font = Font(name='맑은 고딕', size=10, color='FF0000', bold=True)
        ws['A1'].alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 30

        headers = ['이름', '생년월일', '연락처', '학년', '반', '번호']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border

        example_font = Font(name='맑은 고딕', size=10, color='999999')
        examples = ['홍길동', '2008-03-15', '010-1234-5678', '1', '3', '1']
        for col, val in enumerate(examples, 1):
            cell = ws.cell(row=3, column=col, value=val)
            cell.font = example_font; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border

        for row in range(4, 54):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col); cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

        widths = [12, 16, 18, 8, 8, 8]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='학생명단_입력양식.xlsx')

    except Exception as e:
        print(f"템플릿 다운로드 오류: {e}")
        return jsonify({'success': False, 'message': '템플릿 생성 중 오류가 발생했습니다.'})

# ============================================
# 학생 엑셀 업로드 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/upload', methods=['POST'])
def upload_students():
    conn = None
    cursor = None
    try:
        import openpyxl
        file = request.files.get('file')
        school_id = sanitize_input(request.form.get('school_id'), 50)
        member_school = sanitize_input(request.form.get('member_school'), 100)
        class_grade = sanitize_input(request.form.get('class_grade'), 10)
        class_no = sanitize_input(request.form.get('class_no'), 10)

        if not file:
            return jsonify({'success': False, 'message': '파일을 선택해주세요.'})
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Excel 파일(.xlsx)만 업로드 가능합니다.'})

        wb = openpyxl.load_workbook(file)
        ws = wb.active

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        added = 0; updated = 0; skipped = 0; errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or not row[0]: continue
            member_name = str(row[0]).strip() if row[0] else None
            member_birth = str(row[1]).strip() if row[1] and str(row[1]).strip() else None
            member_tel = str(row[2]).strip() if row[2] and str(row[2]).strip() else None
            row_class_grade = str(row[3]).strip() if len(row) > 3 and row[3] else class_grade
            row_class_no = str(row[4]).strip() if len(row) > 4 and row[4] else class_no
            row_class_num = str(row[5]).strip() if len(row) > 5 and row[5] else None

            if not member_name: continue
            if member_name == '홍길동' and member_tel == '010-1234-5678': continue

            try:
                if school_id:
                    cursor.execute("SELECT id, member_birth, member_tel FROM stu_all WHERE school_id = %s AND member_name = %s", (school_id, member_name))
                else:
                    cursor.execute("SELECT id, member_birth, member_tel FROM stu_all WHERE member_school = %s AND member_name = %s", (member_school, member_name))
                candidates = cursor.fetchall()
                existing = None
                for c in candidates:
                    if (member_tel and c['member_tel'] and member_tel == c['member_tel']) or \
                       (member_birth and c['member_birth'] and member_birth == c['member_birth']):
                        existing = c; break

                if existing:
                    update_parts = ["class_grade = %s", "class_no = %s", "updated_at = NOW()"]
                    update_params = [row_class_grade, row_class_no]
                    if row_class_num: update_parts.append("class_num = %s"); update_params.append(row_class_num)
                    if member_tel and not existing['member_tel']: update_parts.append("member_tel = %s"); update_params.append(member_tel)
                    if member_birth and not existing['member_birth']: update_parts.append("member_birth = %s"); update_params.append(member_birth)
                    update_params.append(existing['id'])
                    cursor.execute(f"UPDATE stu_all SET {', '.join(update_parts)} WHERE id = %s", update_params)
                    updated += 1
                else:
                    cursor.execute("""INSERT INTO stu_all (member_id, school_id, member_school, member_name, member_birth, member_tel, class_grade, class_no, class_num, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())""",
                        ('', school_id, member_school, member_name, member_birth, member_tel, row_class_grade, row_class_no, row_class_num))
                    added += 1
            except Exception as row_err:
                errors.append(f'{row_idx}행: {str(row_err)}'); skipped += 1

        conn.commit()
        msg = f'처리 완료: 신규 {added}명, 업데이트 {updated}명'
        if skipped > 0: msg += f', 오류 {skipped}건'
        return jsonify({'success': True, 'message': msg, 'added': added, 'updated': updated, 'skipped': skipped, 'errors': errors[:10]})

    except Exception as e:
        print(f"학생 업로드 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': f'업로드 중 오류가 발생했습니다: {str(e)}'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학생 정보 수정 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/update', methods=['POST'])
def update_student_info():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        student_db_id = data.get('id')
        if not student_db_id:
            return jsonify({'success': False, 'message': 'ID가 필요합니다.'})

        member_name = sanitize_input(data.get('member_name'), 100)
        member_birth = sanitize_input(data.get('member_birth'), 20)
        member_tel = sanitize_input(data.get('member_tel'), 20)
        class_num = sanitize_input(data.get('class_num'), 10)
        class_role = sanitize_input(data.get('class_role'), 30)

        if not member_name:
            return jsonify({'success': False, 'message': '이름은 필수입니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()
        cursor.execute("""
            UPDATE stu_all SET member_name = %s, member_birth = %s, member_tel = %s,
            class_num = %s, class_role = %s, updated_at = NOW() WHERE id = %s
        """, (member_name, member_birth or None, member_tel or None,
              class_num or None, class_role or None, student_db_id))
        conn.commit()
        return jsonify({'success': True, 'message': '학생 정보가 수정되었습니다.'})

    except Exception as e:
        print(f"학생 정보 수정 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '수정 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학부모 정보 수정 API
# ============================================
@homeroom_bp.route('/api/homeroom/parents/update', methods=['POST'])
def update_parent_info():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        parent_db_id = data.get('id')
        if not parent_db_id:
            return jsonify({'success': False, 'message': 'ID가 필요합니다.'})

        member_name = sanitize_input(data.get('member_name'), 100)
        member_tel = sanitize_input(data.get('member_tel'), 20)
        child_name = sanitize_input(data.get('child_name'), 100)
        child_birth = sanitize_input(data.get('child_birth'), 20)

        if not member_name:
            return jsonify({'success': False, 'message': '이름은 필수입니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()
        cursor.execute("""
            UPDATE fm_all SET member_name = %s, member_tel = %s,
            child_name = %s, child_birth = %s, updated_at = NOW() WHERE id = %s
        """, (member_name, member_tel or None, child_name or None,
              child_birth or None, parent_db_id))
        conn.commit()
        return jsonify({'success': True, 'message': '학부모 정보가 수정되었습니다.'})

    except Exception as e:
        print(f"학부모 정보 수정 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '수정 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학급자치 역할 수정 API
# ============================================
@homeroom_bp.route('/api/homeroom/students/role', methods=['POST'])
def update_student_role():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        student_db_id = data.get('id')
        class_role = sanitize_input(data.get('class_role'), 30)

        if not student_db_id:
            return jsonify({'success': False, 'message': 'ID가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()
        cursor.execute("UPDATE stu_all SET class_role = %s, updated_at = NOW() WHERE id = %s",
                        (class_role or None, student_db_id))
        conn.commit()
        return jsonify({'success': True, 'message': '학급자치 역할이 수정되었습니다.'})

    except Exception as e:
        print(f"학급자치 역할 수정 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '역할 수정 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학부모 명단 조회 API
# ============================================
@homeroom_bp.route('/api/homeroom/parents', methods=['GET'])
def get_homeroom_parents():
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        class_grade = sanitize_input(request.args.get('class_grade'), 10)
        class_no = sanitize_input(request.args.get('class_no'), 10)

        if not class_grade or not class_no:
            return jsonify({'success': False, 'message': '학급 정보가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})

        cursor = conn.cursor()

        if school_id:
            cursor.execute("""
                SELECT id, member_id, member_name, member_tel, child_name, child_birth, class_num
                FROM fm_all WHERE school_id = %s AND class_grade = %s AND class_no = %s
                ORDER BY CAST(class_num AS UNSIGNED) ASC, class_num ASC
            """, (school_id, class_grade, class_no))
        else:
            cursor.execute("""
                SELECT id, member_id, member_name, member_tel, child_name, child_birth, class_num
                FROM fm_all WHERE member_school = %s AND class_grade = %s AND class_no = %s
                ORDER BY CAST(class_num AS UNSIGNED) ASC, class_num ASC
            """, (member_school, class_grade, class_no))

        parents = cursor.fetchall()
        parent_list = [{'id': p['id'], 'member_id': p['member_id'], 'member_name': p['member_name'],
                        'member_tel': p['member_tel'] or '',
                        'child_name': p['child_name'] or '',
                        'child_birth': str(p['child_birth']) if p['child_birth'] else '',
                        'class_num': p['class_num'] or ''} for p in parents]

        return jsonify({'success': True, 'parents': parent_list, 'count': len(parent_list)})

    except Exception as e:
        print(f"학부모 명단 조회 오류: {e}")
        return jsonify({'success': False, 'message': '학부모 명단 조회 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학급 공지사항 목록 조회 API
# ============================================
@homeroom_bp.route('/api/homeroom/notice/list', methods=['GET'])
def get_homeroom_notice_list():
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        class_grade = sanitize_input(request.args.get('class_grade'), 10)
        class_no = sanitize_input(request.args.get('class_no'), 10)
        if not class_grade or not class_no:
            return jsonify({'success': False, 'message': '학급 정보가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        # [보안] 교사는 본인 작성 공지만 조회
        teacher_filter = ""
        extra_params = []
        if session.get('user_role') == 'teacher':
            teacher_filter = " AND teacher_id = %s"
            extra_params = [session.get('user_id')]

        # 검색 키워드
        keyword = sanitize_input(request.args.get('keyword'), 100)
        keyword_filter = ""
        keyword_params = []
        if keyword:
            keyword_filter = " AND (title LIKE %s OR content LIKE %s)"
            keyword_params = [f'%{keyword}%', f'%{keyword}%']

        if school_id:
            cursor.execute("SELECT id, teacher_name, title, content, category, created_at FROM homeroom_notice WHERE school_id = %s AND class_grade = %s AND class_no = %s" + teacher_filter + keyword_filter + " ORDER BY created_at DESC LIMIT 50", [school_id, class_grade, class_no] + extra_params + keyword_params)
        else:
            cursor.execute("SELECT id, teacher_name, title, content, category, created_at FROM homeroom_notice WHERE member_school = %s AND class_grade = %s AND class_no = %s" + teacher_filter + keyword_filter + " ORDER BY created_at DESC LIMIT 50", [member_school, class_grade, class_no] + extra_params + keyword_params)

        notices = cursor.fetchall()
        notice_list = [{'id': n['id'], 'teacher_name': n['teacher_name'], 'title': n['title'], 'content': n['content'],
                        'category': n.get('category', '일반') or '일반',
                        'created_at': n['created_at'].strftime('%Y-%m-%d %H:%M') if n['created_at'] else ''} for n in notices]
        return jsonify({'success': True, 'notices': notice_list})

    except Exception as e:
        print(f"학급 공지사항 조회 오류: {e}")
        return jsonify({'success': False, 'message': '공지사항 조회 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학급 공지사항 등록 API
# ============================================
@homeroom_bp.route('/api/homeroom/notice/create', methods=['POST'])
def create_homeroom_notice():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        # [보안] school_id, teacher_id 등은 세션에서 가져옴 (클라이언트 위조 방지)
        school_id = session.get('school_id') or sanitize_input(data.get('school_id'), 50)
        member_school = session.get('user_school') or sanitize_input(data.get('member_school'), 100)
        class_grade = sanitize_input(data.get('class_grade'), 10)
        class_no = sanitize_input(data.get('class_no'), 10)
        teacher_id = session.get('user_id') or sanitize_input(data.get('teacher_id'), 50)
        teacher_name = session.get('user_name') or sanitize_input(data.get('teacher_name'), 100)
        title = sanitize_html(data.get('title', ''), 200)
        content = sanitize_html(data.get('content', ''), 5000)
        category = sanitize_input(data.get('category', '일반'), 20) or '일반'

        if not all([class_grade, class_no, title, content]):
            return jsonify({'success': False, 'message': '필수 항목을 모두 입력해주세요.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        cursor.execute("""INSERT INTO homeroom_notice (school_id, member_school, class_grade, class_no, teacher_id, teacher_name, title, content, category, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())""",
            (school_id, member_school, class_grade, class_no, teacher_id, teacher_name, title, content, category))
        conn.commit()

        # 푸시 알림 발송
        try:
            from utils.push_helper import send_push_to_class
            send_push_to_class(school_id, class_grade, class_no, '학급 공지', title, '/highschool/st_homeroom.html', ['student', 'parent'])
        except Exception as pe:
            print(f"[Homeroom Notice] Push error: {pe}")

        return jsonify({'success': True, 'message': '학급 공지사항이 등록되었습니다.', 'id': cursor.lastrowid})

    except Exception as e:
        print(f"학급 공지사항 등록 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '공지사항 등록 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학급 공지사항 삭제 API
# ============================================
@homeroom_bp.route('/api/homeroom/notice/delete', methods=['POST'])
def delete_homeroom_notice():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        notice_id = sanitize_input(data.get('id'), 20)
        teacher_id = sanitize_input(data.get('teacher_id'), 50)
        if not notice_id:
            return jsonify({'success': False, 'message': '공지사항 ID가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()
        # [보안] 소유자 검증 - 세션 기반
        cursor.execute("SELECT teacher_id FROM homeroom_notice WHERE id = %s", (notice_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'message': '데이터를 찾을 수 없습니다.'}), 404
        if row['teacher_id'] != session.get('user_id'):
            return jsonify({'success': False, 'message': '본인이 작성한 공지사항만 삭제할 수 있습니다.'}), 403
        cursor.execute("DELETE FROM homeroom_notice WHERE id = %s", (notice_id,))
        conn.commit()
        return jsonify({'success': True, 'message': '공지사항이 삭제되었습니다.'})

    except Exception as e:
        print(f"학급 공지사항 삭제 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '삭제 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 학급 공지사항 수정 API
# ============================================
@homeroom_bp.route('/api/homeroom/notice/update', methods=['POST'])
def update_homeroom_notice():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        notice_id = sanitize_input(data.get('id'), 20)
        title = sanitize_html(data.get('title', ''))
        content = sanitize_html(data.get('content', ''))

        if not notice_id:
            return jsonify({'success': False, 'message': '공지사항 ID가 필요합니다.'})
        if not title or not content:
            return jsonify({'success': False, 'message': '제목과 내용을 입력해주세요.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        cursor.execute("SELECT teacher_id FROM homeroom_notice WHERE id = %s", (notice_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'message': '데이터를 찾을 수 없습니다.'}), 404
        if row['teacher_id'] != session.get('user_id'):
            return jsonify({'success': False, 'message': '본인이 작성한 공지사항만 수정할 수 있습니다.'}), 403

        cursor.execute("UPDATE homeroom_notice SET title = %s, content = %s WHERE id = %s",
                        (title, content, notice_id))
        conn.commit()
        return jsonify({'success': True, 'message': '공지사항이 수정되었습니다.'})

    except Exception as e:
        print(f"학급 공지사항 수정 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '수정 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 상담 일정 목록 조회 API
# ============================================
@homeroom_bp.route('/api/homeroom/counsel-schedule/list', methods=['GET'])
def get_counsel_schedule_list():
    # [보안] 교사와 학부모만 조회 가능 (학생 차단)
    user_role = session.get('user_role')
    if user_role not in ('teacher', 'parent'):
        return jsonify({'success': False, 'message': '조회 권한이 없습니다.'}), 403
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        class_grade = sanitize_input(request.args.get('class_grade'), 10)
        class_no = sanitize_input(request.args.get('class_no'), 10)
        student_id = sanitize_input(request.args.get('student_id'), 50)
        # [보안] 학부모는 반드시 student_id 필터 필요
        if user_role == 'parent' and not student_id:
            return jsonify({'success': False, 'message': '자녀 정보가 필요합니다.'})
        if not class_grade or not class_no:
            return jsonify({'success': False, 'message': '학급 정보가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        if school_id:
            query = "SELECT id, student_id, student_name, class_num, counsel_date, counsel_type, memo, status FROM homeroom_counsel_schedule WHERE school_id = %s AND class_grade = %s AND class_no = %s"
            params = [school_id, class_grade, class_no]
        else:
            query = "SELECT id, student_id, student_name, class_num, counsel_date, counsel_type, memo, status FROM homeroom_counsel_schedule WHERE member_school = %s AND class_grade = %s AND class_no = %s"
            params = [member_school, class_grade, class_no]

        # [보안] 교사는 담임 본인만 조회
        if session.get('user_role') == 'teacher':
            query += " AND teacher_id = %s"; params.append(session.get('user_id'))
        if student_id: query += " AND student_id = %s"; params.append(student_id)
        query += " ORDER BY counsel_date DESC"
        cursor.execute(query, params)
        schedules = cursor.fetchall()

        schedule_list = [{'id': s['id'], 'student_id': s['student_id'], 'student_name': s['student_name'],
                          'class_num': s['class_num'], 'counsel_date': s['counsel_date'].strftime('%Y-%m-%d %H:%M') if s['counsel_date'] else '',
                          'counsel_type': s['counsel_type'], 'memo': s['memo'], 'status': s['status']} for s in schedules]
        return jsonify({'success': True, 'schedules': schedule_list})

    except Exception as e:
        print(f"상담 일정 조회 오류: {e}")
        return jsonify({'success': False, 'message': '상담 일정 조회 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 상담 일정 등록 API
# ============================================
@homeroom_bp.route('/api/homeroom/counsel-schedule/create', methods=['POST'])
def create_counsel_schedule():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        # [보안] school_id, member_school은 세션에서 가져옴 (클라이언트 위조 방지)
        school_id = session.get('school_id') or sanitize_input(data.get('school_id'), 50)
        member_school = session.get('user_school') or sanitize_input(data.get('member_school'), 100)
        class_grade = sanitize_input(data.get('class_grade'), 10)
        class_no = sanitize_input(data.get('class_no'), 10)
        # teacher_id: 교사 본인이면 세션, 학부모가 신청하면 요청값 사용
        if session.get('user_role') == 'teacher':
            teacher_id = session.get('user_id')
            teacher_name = session.get('user_name')
        else:
            teacher_id = sanitize_input(data.get('teacher_id'), 50)
            teacher_name = sanitize_input(data.get('teacher_name'), 100)
        student_id = sanitize_input(data.get('student_id'), 50)
        student_name = sanitize_input(data.get('student_name'), 100)
        class_num = sanitize_input(data.get('class_num'), 10)
        counsel_date = sanitize_input(data.get('counsel_date'), 20)
        counsel_type = sanitize_input(data.get('counsel_type'), 50)
        memo = sanitize_html(data.get('memo', ''), 1000)

        if not all([class_grade, class_no, student_name, counsel_date]):
            return jsonify({'success': False, 'message': '필수 항목을 모두 입력해주세요.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        cursor.execute("""INSERT INTO homeroom_counsel_schedule
            (school_id, member_school, class_grade, class_no, teacher_id, teacher_name, student_id, student_name, class_num, counsel_date, counsel_type, memo, status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'scheduled', NOW())""",
            (school_id, member_school, class_grade, class_no, teacher_id, teacher_name, student_id, student_name, class_num, counsel_date, counsel_type, memo))
        conn.commit()

        # 해당 학생+학부모에게 푸시
        if student_id:
            try:
                from utils.push_helper import send_push_to_student
                send_push_to_student(school_id, student_id, '상담 일정 등록', f'{counsel_date} 상담이 예정되었습니다.', '/highschool/st_homeroom.html')
            except Exception as pe:
                print(f"[Counsel Schedule] Push error: {pe}")

        return jsonify({'success': True, 'message': '상담 일정이 등록되었습니다.'})

    except Exception as e:
        print(f"상담 일정 등록 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '상담 일정 등록 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 상담 일정 삭제 API
# ============================================
@homeroom_bp.route('/api/homeroom/counsel-schedule/delete', methods=['POST'])
def delete_counsel_schedule():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        schedule_id = sanitize_input(data.get('id'), 20)
        if not schedule_id:
            return jsonify({'success': False, 'message': 'ID가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()
        # [보안] 소유자 검증
        cursor.execute("SELECT teacher_id FROM homeroom_counsel_schedule WHERE id = %s", (schedule_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'message': '데이터를 찾을 수 없습니다.'}), 404
        if row['teacher_id'] != session.get('user_id'):
            return jsonify({'success': False, 'message': '본인이 작성한 데이터만 삭제할 수 있습니다.'}), 403
        cursor.execute("DELETE FROM homeroom_counsel_schedule WHERE id = %s", (schedule_id,))
        conn.commit()
        return jsonify({'success': True, 'message': '상담 일정이 삭제되었습니다.'})

    except Exception as e:
        print(f"상담 일정 삭제 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '삭제 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 상담 일정 상태 변경 API
# ============================================
@homeroom_bp.route('/api/homeroom/counsel-schedule/update-status', methods=['POST'])
def update_counsel_schedule_status():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        schedule_id = sanitize_input(data.get('id'), 20)
        status = sanitize_input(data.get('status'), 20)
        if not schedule_id or not status:
            return jsonify({'success': False, 'message': 'ID와 상태 값이 필요합니다.'})
        if status not in ('scheduled', 'completed'):
            return jsonify({'success': False, 'message': '유효하지 않은 상태 값입니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()
        cursor.execute("UPDATE homeroom_counsel_schedule SET status = %s WHERE id = %s", (status, schedule_id))
        conn.commit()
        status_text = '완료' if status == 'completed' else '예정'
        return jsonify({'success': True, 'message': f'상담 일정이 {status_text} 상태로 변경되었습니다.'})

    except Exception as e:
        print(f"상담 일정 상태 변경 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '상태 변경 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 상담 일지 목록 조회 API
# ============================================
@homeroom_bp.route('/api/homeroom/counsel-log/list', methods=['GET'])
def get_counsel_log_list():
    # [보안] 교사만 상담 일지 조회 가능
    if session.get('user_role') != 'teacher':
        return jsonify({'success': False, 'message': '교사만 조회할 수 있습니다.'}), 403
    conn = None
    cursor = None
    try:
        school_id = sanitize_input(request.args.get('school_id'), 50)
        member_school = sanitize_input(request.args.get('member_school'), 100)
        class_grade = sanitize_input(request.args.get('class_grade'), 10)
        class_no = sanitize_input(request.args.get('class_no'), 10)
        student_id = sanitize_input(request.args.get('student_id'), 50)
        if not class_grade or not class_no:
            return jsonify({'success': False, 'message': '학급 정보가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        base_cols = "id, student_id, student_name, class_num, counsel_date, counsel_type, content, result, next_plan"
        # [보안] 교사는 담임 본인만 조회
        is_teacher = session.get('user_role') == 'teacher'
        user_id = session.get('user_id')

        # 검색 키워드
        keyword = sanitize_input(request.args.get('keyword'), 100)
        kw_filter = ""
        kw_params = []
        if keyword:
            kw_filter = " AND (content LIKE %s OR result LIKE %s OR student_name LIKE %s)"
            kw_params = [f'%{keyword}%', f'%{keyword}%', f'%{keyword}%']

        if student_id:
            if school_id:
                if is_teacher:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE school_id = %s AND class_grade = %s AND class_no = %s AND teacher_id = %s AND student_id = %s" + kw_filter + " ORDER BY counsel_date DESC", [school_id, class_grade, class_no, user_id, student_id] + kw_params)
                else:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE school_id = %s AND class_grade = %s AND class_no = %s AND student_id = %s" + kw_filter + " ORDER BY counsel_date DESC", [school_id, class_grade, class_no, student_id] + kw_params)
            else:
                if is_teacher:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE member_school = %s AND class_grade = %s AND class_no = %s AND teacher_id = %s AND student_id = %s" + kw_filter + " ORDER BY counsel_date DESC", [member_school, class_grade, class_no, user_id, student_id] + kw_params)
                else:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE member_school = %s AND class_grade = %s AND class_no = %s AND student_id = %s" + kw_filter + " ORDER BY counsel_date DESC", [member_school, class_grade, class_no, student_id] + kw_params)
        else:
            if school_id:
                if is_teacher:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE school_id = %s AND class_grade = %s AND class_no = %s AND teacher_id = %s" + kw_filter + " ORDER BY counsel_date DESC", [school_id, class_grade, class_no, user_id] + kw_params)
                else:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE school_id = %s AND class_grade = %s AND class_no = %s" + kw_filter + " ORDER BY counsel_date DESC", [school_id, class_grade, class_no] + kw_params)
            else:
                if is_teacher:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE member_school = %s AND class_grade = %s AND class_no = %s AND teacher_id = %s" + kw_filter + " ORDER BY counsel_date DESC", [member_school, class_grade, class_no, user_id] + kw_params)
                else:
                    cursor.execute(f"SELECT {base_cols} FROM homeroom_counsel_log WHERE member_school = %s AND class_grade = %s AND class_no = %s" + kw_filter + " ORDER BY counsel_date DESC", [member_school, class_grade, class_no] + kw_params)

        logs = cursor.fetchall()
        log_list = [{'id': l['id'], 'student_id': l['student_id'], 'student_name': l['student_name'],
                     'class_num': l['class_num'], 'counsel_date': l['counsel_date'].strftime('%Y-%m-%d') if l['counsel_date'] else '',
                     'counsel_type': l['counsel_type'], 'content': l['content'], 'result': l['result'], 'next_plan': l['next_plan']} for l in logs]
        return jsonify({'success': True, 'logs': log_list})

    except Exception as e:
        print(f"상담 일지 조회 오류: {e}")
        return jsonify({'success': False, 'message': '상담 일지 조회 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 상담 일지 등록 API
# ============================================
@homeroom_bp.route('/api/homeroom/counsel-log/create', methods=['POST'])
def create_counsel_log():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        # [보안] school_id, teacher_id 등은 세션에서 가져옴 (클라이언트 위조 방지)
        school_id = session.get('school_id') or sanitize_input(data.get('school_id'), 50)
        member_school = session.get('user_school') or sanitize_input(data.get('member_school'), 100)
        class_grade = sanitize_input(data.get('class_grade'), 10)
        class_no = sanitize_input(data.get('class_no'), 10)
        teacher_id = session.get('user_id') or sanitize_input(data.get('teacher_id'), 50)
        teacher_name = session.get('user_name') or sanitize_input(data.get('teacher_name'), 100)
        student_id = sanitize_input(data.get('student_id'), 50)
        student_name = sanitize_input(data.get('student_name'), 100)
        class_num = sanitize_input(data.get('class_num'), 10)
        counsel_date = sanitize_input(data.get('counsel_date'), 20)
        counsel_type = sanitize_input(data.get('counsel_type'), 50)
        content = sanitize_html(data.get('content', ''), 5000)
        result = sanitize_html(data.get('result', ''), 2000)
        next_plan = sanitize_html(data.get('next_plan', ''), 2000)

        if not all([class_grade, class_no, student_name, counsel_date, content]):
            return jsonify({'success': False, 'message': '필수 항목을 모두 입력해주세요.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()

        cursor.execute("""INSERT INTO homeroom_counsel_log
            (school_id, member_school, class_grade, class_no, teacher_id, teacher_name, student_id, student_name, class_num, counsel_date, counsel_type, content, result, next_plan, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())""",
            (school_id, member_school, class_grade, class_no, teacher_id, teacher_name, student_id, student_name, class_num, counsel_date, counsel_type, content, result, next_plan))
        conn.commit()
        return jsonify({'success': True, 'message': '상담 일지가 등록되었습니다.', 'id': cursor.lastrowid})

    except Exception as e:
        print(f"상담 일지 등록 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '상담 일지 등록 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ============================================
# 상담 일지 삭제 API
# ============================================
@homeroom_bp.route('/api/homeroom/counsel-log/delete', methods=['POST'])
def delete_counsel_log():
    conn = None
    cursor = None
    try:
        data = request.get_json()
        log_id = sanitize_input(data.get('id'), 20)
        if not log_id:
            return jsonify({'success': False, 'message': 'ID가 필요합니다.'})

        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': '데이터베이스 연결 오류'})
        cursor = conn.cursor()
        # [보안] 소유자 검증
        cursor.execute("SELECT teacher_id FROM homeroom_counsel_log WHERE id = %s", (log_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'message': '데이터를 찾을 수 없습니다.'}), 404
        if row['teacher_id'] != session.get('user_id'):
            return jsonify({'success': False, 'message': '본인이 작성한 데이터만 삭제할 수 있습니다.'}), 403
        cursor.execute("DELETE FROM homeroom_counsel_log WHERE id = %s", (log_id,))
        conn.commit()
        return jsonify({'success': True, 'message': '상담 일지가 삭제되었습니다.'})

    except Exception as e:
        print(f"상담 일지 삭제 오류: {e}")
        if conn: conn.rollback()
        return jsonify({'success': False, 'message': '삭제 중 오류가 발생했습니다.'})
    finally:
        if cursor: cursor.close()
        if conn: conn.close()
